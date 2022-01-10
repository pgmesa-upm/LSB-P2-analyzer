
# Built-in Modules
import os
import sys
import csv
from time import time as get_time
from math import floor
from pathlib import Path
from shutil import copyfile
from win32com.client import Dispatch

# Dependencies
import numpy as np
from scipy import signal
from scipy.stats import linregress
from tabulate import tabulate
import matplotlib.pyplot as plt
from openpyxl.workbook import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import openpyxl

default_path = '../'; stats_template = 'stats_template.xlsx'; anova_template = 'anova_template.xlsx'
# ------------------------- Custom Settings -------------------------
# -------------------------------------------------------------------
custom_path = None
# Nombre directorios y archivos creados en el analisis
analysis_fname = 'last_analisis.txt'
analysis_dir_name = '__analysis__'
filt_dir_name = 'filtered_data'
norm_dir_name = 'normalized_data'
xcorr_dir_name = 'cross-correlations'
anova_dir_name = 'ANOVA'
# -- Excels
stats_nt_pd_excelname = 'stats_vehiculos_PD.xlsx'
stats_t_pd_excelname = 'stats_tratados_PD.xlsx'
stats_nt_pi_excelname = 'stats_vehiculos_PI.xlsx'
stats_t_pi_excelname = 'stats_tratados_PI.xlsx'
anova_pd_hd_excelname = 'anova_PD_HD.xlsx'
anova_pd_hi_excelname = 'anova_PD_HI.xlsx'
anova_pi_hd_excelname = 'anova_PI_HD.xlsx'
anova_pi_hi_excelname = 'anova_PI_HI.xlsx'
# Nombre directorios ratones
rigth_paw_nt_dir = 'pata-derecha_no-tratados'; rigth_paw_t_dir = 'pata-derecha_tratados'
left_paw_nt_dir = 'pata-izquierda_no-tratados'; left_paw_t_dir = 'pata-izquierda_tratados'
dir_basales = 'basales'; dir_72horas = '72-horas'; dir_2semanas = '2-semanas'
# Stat headers
mouse_header = "Raton"
amplitud_header = "Amplitud mean (uV)"
latency_header = "Latency (ms)"
area_header = "Absolute area"
pearson_header = 'Pearson Correlation'
# -------------------------------------------------------------------
# -------------------------------------------------------------------

if custom_path is not None:
    path = Path(custom_path).resolve()
else:
    path = Path(default_path).resolve()
    
rigth_paw_dirs = [rigth_paw_nt_dir, rigth_paw_t_dir] 
left_paw_dirs = [left_paw_nt_dir, left_paw_t_dir]
paws_dir_names = rigth_paw_dirs + left_paw_dirs
tgroups = [dir_basales, dir_72horas, dir_2semanas]
stat_headers = [mouse_header, amplitud_header, latency_header, area_header, pearson_header]
stats_excel_names = [stats_nt_pd_excelname, stats_t_pd_excelname, stats_nt_pi_excelname, stats_t_pi_excelname]
anova_excelnames = [anova_pd_hd_excelname, anova_pd_hi_excelname, anova_pi_hd_excelname, anova_pi_hi_excelname]
dep_variables = [amplitud_header, latency_header, area_header]

reset_log = True

def timer(func):
    """Mide el tiempo que tarda en ejecutarse una funcion en minutos 
    y segundos

    Args:
        func: funcion a ejecutar
    """
    def f(*a, **ka):
        t0 = get_time()
        func(*a,**ka)
        tf = get_time()
        total_secs = round(tf-t0, 2)
        if total_secs < 60:
            print(f"Elapsed time: {total_secs} s")
        else: 
            mins = floor(total_secs/60)
            secs = int(total_secs - mins*60)
            print(f"Elapsed time: {mins} min {secs} s")
    return f

@timer
def main():
    print(f"[%] Analizando archivos de '{path}'")
   
    analysis_path = path/analysis_dir_name
    if not os.path.exists(analysis_path): os.mkdir(analysis_path)
    if not os.path.exists(analysis_path/norm_dir_name): os.mkdir(analysis_path/norm_dir_name)
    if not os.path.exists(analysis_path/filt_dir_name): os.mkdir(analysis_path/filt_dir_name)
    if not os.path.exists(analysis_path/xcorr_dir_name): os.mkdir(analysis_path/xcorr_dir_name)
    if not os.path.exists(analysis_path/anova_dir_name): os.mkdir(analysis_path/anova_dir_name)
    for excel in stats_excel_names:
        if os.path.exists(analysis_path/excel): 
            os.remove(analysis_path/excel)
    for excel in anova_excelnames:
        if os.path.exists(analysis_path/anova_dir_name/excel): 
            os.remove(analysis_path/anova_dir_name/excel)
    
    log(f" + Analysing '{path}'")
    letters = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for ipaw, paw in enumerate(paws_dir_names):
        log(f" => {paw}")
        gcounter = 0
        for grp in tgroups:
            log(f" -> {grp}")
            grp_path = path/paw/grp
            files = [name for name in os.listdir(grp_path) if name.endswith('.txt')]
            headers = []
            for i, h in enumerate(stat_headers): 
                if i == 0: headers.append(h); continue
                if h == pearson_header:
                    if grp == dir_basales:
                        headers.append(h)
                    continue
                headers.append(h + " HD")
                headers.append(h + " HI")   
                
            table = [headers]
            fcounter = 0
            for ifile, fname in enumerate(files):
                print(f"Analizando => '{fname}'...")
                t = []; hd = []; hi = []; txt_headers = [] 
                with open(grp_path/fname, 'r') as file:
                    reader = csv.reader(file, delimiter='\t')
                    for i, line in enumerate(reader):
                        if i == 0: txt_headers.append(line); continue
                        if line[0] == "" or line[1] == "" or line[2] == "": continue
                        t.append(float(line[0])); hd.append(float(line[1])); hi.append(float(line[2]))
                print("   - Filtrando...")
                t, hd, hi = filt(t, hd, hi)
                assert len(t) == len(hd) and len(hd) == len(hi)
                # Guardamos datos filtrados (quitamos los ultimos 11 ms)
                save_data(t, hd, hi, analysis_path/filt_dir_name, f"{paw}/{grp}/{(fname.replace('.', '_filt.'))}", headers=txt_headers)
                print(f"   - Calculando {stat_headers}...")
                # HD
                hd_stats = calc_stats(t, hd)
                if '-s' in sys.argv:
                    plt.plot(t, hd)
                # HI
                hi_stats = calc_stats(t, hi)
                if '-s' in sys.argv:
                    plt.plot(t, hi)
                # all
                stats = [fname]
                for hd_stat, hi_stat in zip(hd_stats, hi_stats):
                    stats.append(hd_stat)
                    stats.append(hi_stat)
                if grp == dir_basales:
                    # Calculamos correlacion de Pearson
                    print("   - Calculando Correlacion de Pearson (R y P)...")
                    rval, pval = linregress(hd,hi)[2:4]
                    rval = round(rval, 8); pval = round(pval, 8) 
                    # Si pval<0.05 significa que el R calculado es suficientemente distinto de 0
                    # para decir que existe relacion lineal entre las señales
                    pearson = f"    R={rval} | P={pval}"
                    stats.append(pearson)
                table.append(stats)
                print(f"   - Actualizando excel...")
                # Updateamos el excel
                excel_path = analysis_path/stats_excel_names[ipaw]
                if not os.path.exists(excel_path):
                    copyfile(stats_template, excel_path)
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
                offset = 4; counter = fcounter + gcounter
                for i, stat in enumerate(stats):
                    if i == 0: continue
                    letter = letters[i-1]
                    cell = letter+str(counter+offset)
                    if letter == 'J': 
                        stat = rval
                        ws[f'K{counter+offset}'] = pval
                    ws[cell] = stat
                if grp == dir_basales:
                    # Realizamos la correlacion cruzada entre hemisferios
                    print(f"   - Realizando Correlacion Cruzada...")
                    # Normalizamos los datos en el rango [-1, 1]
                    norm_hd = hd/np.linalg.norm(hd); norm_hi = hi/np.linalg.norm(hi)
                    assert max(norm_hd) <= 1 and min(norm_hi) >= -1 
                    save_data(t, norm_hd, norm_hi, analysis_path/norm_dir_name, f"{paw}/{grp}/{(fname.replace('.', '_norm.'))}", headers=txt_headers)
                    chart_ws = Workbook.create_sheet(wb, title=f'Correlation-{ifile+1}')
                    xcorr = list(signal.correlate(norm_hd, norm_hi, mode='same'))
                    chart_ws.append(["Lags", "Cross Correlation"])
                    lags = signal.correlation_lags(norm_hi.size, norm_hd.size, mode="same")
                    # Guardmos la imagen
                    plt.plot(lags, xcorr)
                    plt.grid()
                    plt.xlabel("Lags"); plt.ylabel("Cross Correlation")         
                    save_graph(analysis_path/xcorr_dir_name, f"{paw}/{grp}/{(fname.replace('.txt', '_xcorr.png'))}")
                    plt.clf()
                    for lag, val in zip(lags, xcorr):
                        chart_ws.append([lag, val])
                    xvalues = Reference(chart_ws, min_col=1, min_row=2, max_row=len(lags))
                    values = Reference(chart_ws, min_col=2, min_row=2, max_row=len(xcorr))
                    series = Series(values, xvalues)
                    chart = ScatterChart()
                    chart.title = f"Correlacion Cruzada '{fname}'"
                    chart.x_axis.title = "Lags"
                    chart.y_axis.title = "Cross Correlation"
                    chart.series.append(series)
                    chart_ws.add_chart(chart, 'D2')
                wb.save(path/excel_path)
                # Argumento para plotear
                if '-s' in sys.argv:
                    print(f"   - Mostrando Grafica...")
                    plt.show()
                fcounter += 3
            gcounter += 1
            log(tabulate(table, headers='firstrow', tablefmt='fancy_grid', numalign='center'))
    print("[%] Calculando estadísticos ANOVA...")
    # Calculamos estadístico ANOVA (rellenamos datos en la plantilla)
    anova_path = analysis_path/anova_dir_name
    for i, name in enumerate(anova_excelnames):
        anova_excel_path = anova_path/name
        print(f"[%] Creando '{name}'")
        if not os.path.exists(anova_excel_path):
            copyfile(anova_template, anova_excel_path)
        # Cargamos las hojas excel con las stats calculadas
        index1 = 0; index2 = 1; h_offset = 0
        if i >= 1:
            index1 = 2; index2 = 3
        nt_excel_name = stats_excel_names[index1]
        t_excel_name = stats_excel_names[index2]
        print(nt_excel_name, t_excel_name)
        nt_excel_path = analysis_path/nt_excel_name
        t_excel_path = analysis_path/t_excel_name
        if i%2 != 0:
            h_offset = 1
        refresh_formulas(t_excel_path)
        refresh_formulas(nt_excel_path)
        nt_excel = openpyxl.load_workbook(nt_excel_path, data_only=True)
        nt_ws = nt_excel['stats']
        t_excel = openpyxl.load_workbook(t_excel_path, data_only=True)
        t_ws = t_excel['stats']
        start = 16; offset = 6
        anova_wb = openpyxl.load_workbook(anova_excel_path)
        # Guardamos los datos en el excel de ANOVA 
        anova_letters = ['F', 'G', 'H']; anova_offset = 5; anova_start = 5
        for j, dep_h in enumerate(dep_variables):
            anova_ws = anova_wb[dep_h]
            for k in range(3):
                letter = letters[(j*2)+h_offset]
                nt_mean = nt_ws[letter+str(start+k)].value
                nt_variance = nt_ws[letter+str(start+k+offset)].value
                t_mean = t_ws[letter+str(start+k)].value
                t_variance = t_ws[letter+str(start+k+offset)].value
                anova_letter = anova_letters[k]
                anova_ws[anova_letter+str(anova_start)] = nt_mean
                anova_ws[anova_letter+str(anova_start+1)] = t_mean
                anova_ws[anova_letter+str(anova_start+anova_offset)] = nt_variance
                anova_ws[anova_letter+str(anova_start+anova_offset+1)] = t_variance
        anova_wb.save(anova_excel_path)
    print(f"[%] Se ha creado '/{analysis_dir_name}'")
    print("[%] Fin del analisis")
    
def log(msg:str):
    global reset_log
    mode = 'ab'
    if reset_log:
        mode = 'wb'; reset_log = False
    with open(path/analysis_dir_name/analysis_fname, mode) as file:
        file.write((msg+"\n").encode('utf-8'))
        
def create_path(dir_base, fname):
    dirs = fname.split("/")
    fname = dirs.pop(len(dirs)-1)
    
    for dir_ in dirs:
        if not os.path.exists(dir_base/dir_):
            os.mkdir(dir_base/dir_)
        dir_base = dir_base/dir_
        
    file_path = dir_base/fname

    return file_path

def save_data(t, hd, hi, dir_base:Path, fname:str, headers:list=None):
    file_path = create_path(dir_base, fname)
    
    with open(file_path, 'w') as file:
        writer = csv.writer(file, delimiter='\t')
        if headers is not None: writer.writerow(*headers)
        for t_val, hd_val, hi_val in zip(t, hd, hi):
            writer.writerow([t_val, hd_val, hi_val])
        
def save_graph(dir_base:Path, fname:str):
    file_path = create_path(dir_base, fname)
    plt.savefig(file_path)
    
def refresh_formulas(excel_path):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(excel_path)
    xlBook.RefreshAll()
    xlBook.Save()
    xlBook.Close()

def filt(t, hd, hi, filt_time_ms=10):
    end_index_limit = None;
    for i, time in enumerate(t):
        if time >= filt_time_ms:
            end_index_limit = i
            break         
    return t[end_index_limit:], hd[end_index_limit:], hi[end_index_limit:]    
    
def calc_stats(t:list, array:list) -> list:
    stats = []
    # Amplitud media
    mean = np.mean(array)
    stats.append(round(mean,8))
    # Latencia
    # -- Realizamos un primer filtrado (hasta que la señal no pase de crecer/decrecer 
    # al contrario una vez, no empieza a contar)
    last_val = None; valid_t = None; valid_array = None; tendency = None
    for i, val in enumerate(np.abs(array)):
        if i == 0: last_val = array[0]; continue
        if last_val > val:
            tendency_val = 'decreasing'
        elif last_val < val:
            tendency_val = 'increasing'
        else:
            continue
        if tendency is None:
            tendency = tendency_val
        elif tendency_val != tendency:
            valid_t = t[i:]; valid_array = array[i:]
            break
        last_val = val 
    max_ = max(valid_array); min_ = min(valid_array)
    if abs(max_) >= abs(min_):
        latency = valid_t[valid_array.index(max_)]
    else:
        latency = valid_t[valid_array.index(min_)]
    stats.append(round(latency,8))
    # Area absoluta
    area = np.trapz(np.abs(array), dx=abs(t[1]-t[0]))
    stats.append(round(area,8))
    
    return stats

if __name__ == "__main__":
    try:
        print("[%] Program started")
        main()
        print("[%] Program finished successfully")
    except KeyboardInterrupt:
        print("[!] Exit")
        exit(1)
    # except Exception as err:
    #     print(f"[!] Unexpected Error: {err}")